import os
from datetime import datetime
from typing import Any

from openpyxl import Workbook

from ventilacia_ai.services.config_service import REPORTS_FOLDER


def create_excel_file(results: list[dict[str, Any]], filename: str | None = None) -> str:
    """Создаёт Excel файл: Код | Наименование полное | Ед.изм | Количество (без заголовков)."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"результат_сопоставления_{timestamp}.xlsx"

    filepath = os.path.join(REPORTS_FOLDER, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "Сопоставление"

    for row_idx, result in enumerate(results, 1):
        ws.cell(row=row_idx, column=1, value=result.get("matched_code") or "")
        ws.cell(row=row_idx, column=2, value=result.get("matched_name") or "")
        ws.cell(row=row_idx, column=3, value=result.get("matched_unit") or "")
        ws.cell(row=row_idx, column=4, value=result.get("quantity", ""))

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(filepath)
    return filename
